from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

# シートを追加
ws2 = wb.create_sheet()

for i in range(1, 11):
    if i == 1:
        for j in range(1, 11):
            ws2.cell(row=i, column=j).value = j
    else:
        for j in range(1, 11):
            ws2.cell(row=i, column=j).value = i + j * 10

wb.save("sample.xlsx")